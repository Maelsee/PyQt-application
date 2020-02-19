import json

import openpyxl
from pyecharts import Graph


class hitIterator():
    import numpy as np

    __doc__ = '''计算图中Authority,hub '''
    def __init__(self, edges, path="",):
        self.edges = edges
        self.nodes = []
        self.path = path

    def hits(self):
        for edge in self.edges:
            if edge[0] not in self.nodes:
                self.nodes.append(edge[0])
            if edge[1] not in self.nodes:
                self.nodes.append(edge[1])
        # print(self.nodes)

        N = len(self.nodes)

        # 将节点符号（字母），映射成阿拉伯数字，便于后面生成S矩阵
        i = 0
        node_to_num = {}
        for node in self.nodes:
            node_to_num[node] = i
            i += 1
        for edge in self.edges:
            edge[0] = node_to_num[edge[0]]
            edge[1] = node_to_num[edge[1]]
        # print(self.edges)
        # 生成初步的S矩阵
        S = self.np.zeros([N, N])
        for edge in self.edges:
            S[edge[0], edge[1]] = round(edge[2], 2)
        # print(S)

        # 计算比例：即进行列的归一化处理
        for i in range(N):
            sum_of_row = 0  # = sum(S[:,j])
            for edge in self.edges:
                if edge[0] == i:
                    sum_of_row += round(edge[2], 2)

            if sum_of_row != 0:
                for j in range(N):
                    S[i, j] /= sum_of_row

        S_trans = self.np.transpose(S)  # 转置矩阵
        # print(S)
        # print(S_trans)
        H_matrix_old = self.np.ones((N, 1))
        A_matrix_old = self.np.ones((N, 1))

        e = 100000  # 误差初始化
        k = 0  # 记录迭代次数
        print('hitsloop...')

        while e > 0.00000001:  # 开始迭代
            A_matrix_new = self.np.dot(S_trans, H_matrix_old)  # 迭代公式
            H_matrix_new = self.np.dot(S, A_matrix_new)  # 迭代公式

            a = A_matrix_new ** 2
            a = self.np.sqrt(a.sum())
            A_matrix_new /= a

            h = H_matrix_new ** 2
            h = self.np.sqrt(h.sum())
            H_matrix_new /= h

            e1 = A_matrix_new - A_matrix_old
            e1 = max(map(abs, e1))  # 计算误差
            e2 = H_matrix_new - H_matrix_old
            e2 = max(map(abs, e2))
            e = e1 + e2

            H_matrix_old = H_matrix_new
            A_matrix_old = A_matrix_new
            k += 1

        H_matrixList = H_matrix_old.tolist()
        Hubmax = self.np.max(H_matrixList)
        node_to_Hub = {}

        for i in range(len(H_matrixList)):
            for key, value in node_to_num.items():
                if value == i:
                    node_to_Hub[key] = H_matrixList[i][0]
                    break
        Hub_sortedlist = sorted(node_to_Hub.items(), key=lambda x: x[1], reverse=True)



        A_matrixList = A_matrix_old.tolist()
        Autmax = self.np.max(A_matrixList)
        node_to_Aut = {}
        for i in range(len(A_matrixList)):
            for key, value in node_to_num.items():
                if value == i:
                    node_to_Aut[key] = A_matrixList[i][0]
                    break
        Aut_sortedlist = sorted(node_to_Aut.items(), key=lambda x: x[1], reverse=True)

        edgess = []
        for i in Aut_sortedlist:
            edgess.append([i[0],i[1],node_to_Hub[i[0]]])


        for i in self.edges:
            for key, value in node_to_num.items():
                if value == i[0]:
                    i[0] = key
                if value == i[1]:
                    i[1] = key



        self.echartsshow(self.path,Aut_sortedlist,10)
        self.writeexecl(self.path,edgess)
        print("hits运行完成")
        return True

    def echartsshow(self,filepath,sortedList,C):
        per10nodes = []
        edgesList=[]
        nodesList=[]
        allnodes=set()

        for i in range(C):
           node=sortedList[i][0]
           per10nodes.append(node)
           for x in self.edges:
                if x[0]==node:
                    edgesList.append({"source": node, "target": x[1], "value": x[2]})
                    allnodes.add(x[0])
                    allnodes.add(x[1])
                if x[1]==node:
                    edgesList.append({"source": x[0], "target": node, "value": x[2]})
                    allnodes.add(x[0])
                    allnodes.add(x[1])

        for i in allnodes:
            if i in per10nodes:
                nodesList.append({"name": i, "symbol": "triangle", "symbolSize": 30})
            else:
                nodesList.append({"name": i, "symbol": "circle", "symbolSize": 12})
        nodes=nodesList
        links=edgesList
        graph = Graph("Aut值前"+str(C)+"图例", width=1920, height=1080)
        graph.add(
            "",
            nodes,
            links,
            is_label_show=False,
            repulsion=50,
            is_focusnode=True,
            is_roam=True,
            graph_layout='force',
            line_color="rgba（50,50,50,0.7）",
            graph_edge_symbol=['circle', 'arrow']
        )
        graph.render(path=filepath + '/Hitsresult.html')


    def writeexecl(self,filepath,rows):
        wb = openpyxl.Workbook()

        try:
           for i in wb.sheetnames:
               del wb[i]
        except KeyError:
            pass

        my_sheet = wb.create_sheet()
        nodes=["账户", 'aut值',"hub值"]
        my_sheet.append(nodes)
        for i in rows:
            my_sheet.append(i)

        wb.save(filepath+"/Hitsresult.xlsx")


# if __name__ == '__main__':
#     with open(r"test.json","r",encoding="utf-8") as f:
#         edges = json.load(f)["edges"]
#     hit = hitIterator(edges,path="/home/dang/Nutstore Files/我的坚果云/GUI/project")
#
#     hit.hits()
