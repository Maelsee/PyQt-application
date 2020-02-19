import xlrd
import openpyxl
import json
import os

from pyecharts import Graph

from   DegreeMembership  import  DegreeMembership

def filetomembership(execlpath):
    fileFormat = os.path.splitext(execlpath)[1]
    print(fileFormat)
    if fileFormat == '.xls':
        xlsVisual(execlpath)
        return "支持此格式文件"
    elif fileFormat == '.xlsx':
        xlsxVisual(execlpath)
        return "支持此格式文件"
    elif fileFormat == '.json':
        jsonVisual(execlpath)
        return "支持此格式文件"
    else:
        print("不支持此格式文件！")
        return "不支持此格式文件！"


# xls
def xlsVisual(execlpath, sourceaccount='交易账卡号', targetaccount='对手账号', money='交易金额', label='收付标志'):
    linksList = []
    workbook = xlrd.open_workbook(execlpath)
    sheets = workbook.sheet_names()
    for sheetnum in sheets:
        worksheet = workbook.sheet_by_name(sheetnum)
        firstrowvalue = worksheet.row_values(0)
        print(firstrowvalue)

        # 查找原账户位置
        if sourceaccount in firstrowvalue:
            sourceIndex = firstrowvalue.index(sourceaccount)
        else:
            print("文件中没有%s属性", sourceaccount)
            break
        # 查找目的账户位置
        if targetaccount in firstrowvalue:
            targetIndex = firstrowvalue.index(targetaccount)
        else:
            print("文件中没有%s属性", targetaccount)
            break
            # 查找权值位置
        if money in firstrowvalue:
            moneyIndex = firstrowvalue.index(money)
        else:
            print("文件中没有%s属性", money)
            break
        if label not in firstrowvalue:
            for i in range(1, worksheet.nrows):
                onerowdict = {}
                sourcenode = worksheet.cell_value(i, sourceIndex)
                targetnode = worksheet.cell_value(i, targetIndex)
                moneynode = float(worksheet.cell_value(i, moneyIndex))

                if sourcenode == '' or targetnode == '':
                    continue

                if len(linksList) > 0:
                    flog = 0
                    for x in linksList:
                        if sourcenode == x['source'] and targetnode == x['target']:
                            x['money'] += moneynode
                            x['time'] += 1

                            flog = 1
                    if flog == 0:
                        onerowdict['source'] = sourcenode
                        onerowdict['target'] = targetnode
                        onerowdict['money'] = moneynode
                        onerowdict['time'] = 0
                        linksList.append(onerowdict)
                else:
                    onerowdict['source'] = sourcenode
                    onerowdict['target'] = targetnode
                    onerowdict['money'] = moneynode
                    onerowdict['time'] = 0
                    linksList.append(onerowdict)
        else:
            labelIndex = firstrowvalue.index(label)
            for i in range(1, worksheet.nrows):
                onerowdict = {}
                sourcenode = worksheet.cell_value(i, sourceIndex)
                targetnode = worksheet.cell_value(i, targetIndex)
                moneynode = float(worksheet.cell_value(i, moneyIndex))

                if sourcenode == '' or targetnode == '':
                    continue

                if worksheet.cell_value(i, labelIndex) == '出':

                    if len(linksList) > 0:
                        flog = 0
                        for x in linksList:
                            if sourcenode == x['source'] and targetnode == x['target']:
                                x['money'] += moneynode
                                if moneynode>0:
                                    x['time'] += 1
                                flog = 1
                        if flog == 0:
                            onerowdict['source'] = sourcenode
                            onerowdict['target'] = targetnode
                            onerowdict['money'] = moneynode
                            onerowdict['time'] = 0
                            linksList.append(onerowdict)
                    else:
                        onerowdict['source'] = sourcenode
                        onerowdict['target'] = targetnode
                        onerowdict['money'] = moneynode
                        onerowdict['time'] = 0
                        linksList.append(onerowdict)

                if worksheet.cell_value(i, labelIndex) == '进':

                    if len(linksList) > 0:
                        flog = 0
                        for x in linksList:
                            if targetnode == x['source'] and sourcenode == x['target']:
                                x['money'] += moneynode
                                if moneynode > 0:
                                    x['time'] += 1
                                flog = 1
                        if flog == 0:
                            onerowdict['source'] = targetnode
                            onerowdict['target'] = sourcenode
                            onerowdict['money'] = moneynode
                            onerowdict['time'] = 0
                            linksList.append(onerowdict)
                    else:
                        onerowdict['source'] = targetnode
                        onerowdict['target'] = sourcenode
                        onerowdict['money'] = moneynode
                        onerowdict['time'] = 0
                        linksList.append(onerowdict)
    filepath = os.path.dirname(execlpath)
    print(filepath)
    edges = []

    De=DegreeMembership()
    for x in linksList:
        value = De.membershipFun(money=x['money'],times= x['time'])
        edges.append([x['source'], x['target'], value])

    writeexecl(filepath, edges)
    writejson(filepath, edges)
    writehtml(filepath, edges)

# xlsx
def xlsxVisual(execlpath, sourceaccount='交易账卡号', targetaccount='对手账号', money='交易金额', label='收付标志'):
    linksList = []
    firstrowvalue = []
    workbook = openpyxl.load_workbook(execlpath)
    sheets = workbook.sheetnames
    for sheetnum in sheets:
        worksheet = workbook[sheetnum]
        for cell in list(worksheet.rows)[0]:
            firstrowvalue.append(cell.value)
        print(firstrowvalue)
        # 查找原账户位置
        if sourceaccount in firstrowvalue:
            sourceIndex = firstrowvalue.index(sourceaccount) + 1
        else:
            print("文件中没有%s属性", sourceaccount)
            break
        # 查找目的账户位置
        if targetaccount in firstrowvalue:
            targetIndex = firstrowvalue.index(targetaccount) + 1
        else:
            print("文件中没有%s属性", targetaccount)
            break
            # 查找权值位置
        if money in firstrowvalue:
            moneyIndex = firstrowvalue.index(money) + 1
        else:
            print("文件中没有%s属性", money)
            break
        if label in firstrowvalue:

            for i in range(2, worksheet.max_row):
                onerowdict = {}
                sourcenode = worksheet.cell(row=i, column=sourceIndex).value
                targetnode = worksheet.cell(row=i, column=targetIndex).value
                moneynode = float(worksheet.cell(row=i, column=moneyIndex).value)

                if sourcenode == '' or targetnode == '' or sourcenode == None or targetnode == None:
                    continue
                if len(linksList) > 0:
                    flog = 0
                    for x in linksList:
                        if sourcenode == x['source'] and targetnode == x['target']:
                            x['money'] += moneynode
                            if moneynode>0:
                                x["time"]+=1
                            flog = 1
                    if flog == 0:
                        onerowdict['source'] = sourcenode
                        onerowdict['target'] = targetnode
                        onerowdict['money'] = moneynode
                        onerowdict['time'] = 0
                        linksList.append(onerowdict)
                else:
                    onerowdict['source'] = sourcenode
                    onerowdict['target'] = targetnode
                    onerowdict['money'] = moneynode
                    onerowdict['time'] = 0
                    linksList.append(onerowdict)

        else:
            labelIndex=firstrowvalue.index(label) + 1

            for i in range(2, worksheet.max_row):
                onerowdict = {}
                sourcenode = worksheet.cell(row=i, column=sourceIndex).value
                targetnode = worksheet.cell(row=i, column=targetIndex).value
                moneynode = float(worksheet.cell(row=i, column=moneyIndex).value)


                if sourcenode == '' or targetnode == '' or sourcenode == None or targetnode == None:
                    continue

                if worksheet.cell(row=i, column=labelIndex).value == '出':

                    if len(linksList) > 0:
                        flog = 0
                        for x in linksList:
                            if sourcenode == x['source'] and targetnode == x['target']:
                                x['money'] += moneynode
                                if moneynode > 0:
                                    x["time"] += 1
                                flog = 1
                        if flog == 0:
                            onerowdict['source'] = sourcenode
                            onerowdict['target'] = targetnode
                            onerowdict['money'] = moneynode
                            onerowdict['time'] = 0
                            linksList.append(onerowdict)
                    else:
                        onerowdict['source'] = sourcenode
                        onerowdict['target'] = targetnode
                        onerowdict['money'] = moneynode
                        onerowdict['time'] = 0
                        linksList.append(onerowdict)
                if worksheet.cell(row=i, column=labelIndex).value == '进':

                    if len(linksList) > 0:
                        flog = 0
                        for x in linksList:
                            if targetnode == x['source'] and sourcenode == x['target']:
                                x['money'] += moneynode
                                if moneynode > 0:
                                    x["time"] += 1
                                flog = 1
                        if flog == 0:
                            onerowdict['source'] = targetnode
                            onerowdict['target'] = sourcenode
                            onerowdict['money'] = moneynode
                            onerowdict['time'] = 0
                            linksList.append(onerowdict)
                    else:
                        onerowdict['source'] = targetnode
                        onerowdict['target'] = sourcenode
                        onerowdict['money'] = moneynode
                        onerowdict['time'] = 0
                        linksList.append(onerowdict)
        filepath = os.path.dirname(execlpath)
        edges=[]
        degmembership=DegreeMembership()
        for x in linksList:
            value=degmembership.membershipFun(x['money'],x['time'])
            edges.append([x['source'],x['target'],value])

        writeexecl(filepath,edges)
        writejson(filepath, edges)
        writehtml(filepath, edges)


def jsonVisual(path1):
    with open(path1,"r",encoding="utf-8") as f:
        filedict=json.load(f)

    filepath=os.path.dirname(path1)

    # with open(filepath+'/membership.json',"w+",encoding="utf-8") as f:
    #     json.dump(filedict,f)

    writeexecl(filepath, filedict["edges"])
    writejson(filepath, filedict["edges"])
    writehtml(filepath, filedict["edges"])

def writeexecl(filepath,rows):
    wb = openpyxl.Workbook()

    try:
       for i in wb.sheetnames:
           del wb[i]
    except KeyError:
        pass

    my_sheet = wb.create_sheet()
    edges=["源账户", '对手账户', '隶属度']
    my_sheet.append(edges)
    for i in rows:
        my_sheet.append(i)

    wb.save(filepath+"/membership.xlsx")

def writejson(filepath,rows):
    with open(filepath+'/membership.json',"w+",encoding="utf-8") as f:
        json.dump({"edges":rows},f)

def writehtml(filepath,rows):
    print(rows)
    nodes=[]
    links=[]
    nodeSet=set()
    for i in rows:
        nodeSet.add(i[0])
        nodeSet.add(i[1])

        links.append({"source": i[0], "target": i[1], "value": i[2]})

    for i in nodeSet:
        nodes.append({"name": i, "symbol": "circle", "symbolSize": 12})


        graph = Graph("隶属度" + "图例", width=1920, height=1080)
        graph.add(
            "",
            nodes,
            links,
            is_label_show=False,
            repulsion=50,
            is_focusnode=True,
            is_roam=True,
            graph_layout='force',
            line_color="rgba（50,50,50,0.7）",
            graph_edge_symbol=['circle', 'arrow']
        )
        graph.render(path=filepath + '/membership.html')




# if __name__ == "__main__":
#     execlpath = r'/home/dang/Nutstore Files/我的坚果云/GUI/project/账户交易明细表.xls'
#     execlpath1=r'/home/dang/Nutstore Files/我的坚果云/GUI/project/test.json'
#     filetomembership(execlpath)
#     jsonVisual(execlpath1)

