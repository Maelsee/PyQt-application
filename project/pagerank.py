import json
import os
from pyecharts import  Graph
import openpyxl

class pagerankIterator:
    import numpy as np

    __doc__ = '''计算图中pagerank '''

    def __init__(self, edges, path="/"):
        self.edges = edges
        self.nodes = []
        self.path = path



    def pagerank(self):
        for edge in self.edges:
            if edge[0] not in self.nodes:
                self.nodes.append(edge[0])
            if edge[1] not in self.nodes:
                self.nodes.append(edge[1])
        # print(self.nodes)

        N = len(self.nodes)

        # 将节点符号（字母），映射成阿拉伯数字，便于后面生成A矩阵/S矩阵
        i = 0
        node_to_num = {}
        for node in self.nodes:
            node_to_num[node] = i
            i += 1
        for edge in self.edges:
            edge[0] = node_to_num[edge[0]]
            edge[1] = node_to_num[edge[1]]
        # print(self.edges)

        # 生成初步的S矩阵
        S = self.np.zeros([N, N])
        for edge in self.edges:
            S[edge[0], edge[1]] = round(edge[2], 2)
        # print(S)

        # 计算比例：即一个网页对其他网页的PageRank值的贡献，即进行列的归一化处理
        for i in range(N):
            sum_of_row = 0  # = sum(S[:,j])
            for edge in self.edges:
                if edge[0] == i:
                    sum_of_row += round(edge[2], 2)

            if sum_of_row != 0:
                for j in range(N):
                    S[i, j] /= sum_of_row

        S = self.np.transpose(S)  # 转置矩阵
        # print(S)

        # 计算矩阵A
        alpha = 0.85
        A = alpha * S + (1 - alpha) / N * self.np.ones([N, N])
        # print(A)

        # 生成初始的PageRank值，记录在P_n中，P_n和P_n1均用于迭代
        P_n = self.np.ones((N, 1))
        P_n1 = self.np.zeros((N, 1))

        e = 100000  # 误差初始化
        k = 0  # 记录迭代次数
        print('pagerankloop...')

        while e > 0.00000001:  # 开始迭代
            P_n1 = self.np.dot(A, P_n)  # 迭代公式
            e = P_n1 - P_n
            e = max(map(abs, e))  # 计算误差
            P_n = P_n1
            k += 1
            # print('iteration %s:'%str(k), P_n1)

        P_nList = P_n.tolist()
        PRmax = self.np.max(P_n)
        node_to_pr = {}

        for i in range(len(P_nList)):
            for key, value in node_to_num.items():
                if value == i:
                    node_to_pr[key] = P_nList[i][0]
                    break
        sortedlist = sorted(node_to_pr.items(),
                            key=lambda x: x[1], reverse=True)

        for i in self.edges:
            for key, value in node_to_num.items():
                if value == i[0]:
                    i[0]=key
                if value ==i[1]:
                    i[1]=key

        print(self.edges)

        self.writeexecl(self.path,sortedlist)
        self.echartsshow(self.path,sortedlist,10)
        print("pagerank完成")
        return True




    def echartsshow(self,filepath,sortedList,C):
        per10nodes = []
        edgesList=[]
        nodesList=[]
        allnodes=set()

        for i in range(C):
           node=sortedList[i][0]
           per10nodes.append(node)
           for x in self.edges:
                if x[0]==node:
                    edgesList.append({"source": node, "target": x[1], "value": x[2]})
                    allnodes.add(x[0])
                    allnodes.add(x[1])
                if x[1]==node:
                    edgesList.append({"source": x[0], "target": node, "value": x[2]})
                    allnodes.add(x[0])
                    allnodes.add(x[1])

        for i in allnodes:
            if i in per10nodes:
                nodesList.append({"name": i, "symbol": "triangle", "symbolSize": 30})
            else:
                nodesList.append({"name": i, "symbol": "circle", "symbolSize": 12})
        nodes=nodesList
        links=edgesList
        graph = Graph("pr值前"+str(C)+"图例", width=1920, height=1080)
        graph.add(
            "",
            nodes,
            links,
            is_label_show=False,
            repulsion=50,
            is_focusnode=True,
            is_roam=True,
            graph_layout='force',
            line_color="rgba（50,50,50,0.7）",
            graph_edge_symbol=['circle', 'arrow']
        )
        graph.render(path=filepath + '/Prresult.html')


    def writeexecl(self,filepath,rows):
        wb = openpyxl.Workbook()

        try:
           for i in wb.sheetnames:
               del wb[i]
        except KeyError:
            pass

        my_sheet = wb.create_sheet()
        nodes=["账户", 'pr值']
        my_sheet.append(nodes)
        for i in rows:
            my_sheet.append(i)

        wb.save(filepath+"/Prresult.xlsx")

# if __name__ == '__main__':
#     with open(r"test.json","r",encoding="utf-8") as f:
#         edges = json.load(f)["edges"]
#
#     hit = pagerankIterator(edges,path="/home/dang/Nutstore Files/我的坚果云/GUI/project")
#
#     hit.pagerank()
